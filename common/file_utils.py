"""
文件操作工具类
支持读取YAML/Excel/JSON等格式的测试数据
"""
import os
import json
import yaml
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None
    load_workbook = None


class FileUtils:
    """文件操作工具类"""
    
    @staticmethod
    def read_yaml(file_path: str) -> Dict[str, Any]:
        """
        读取YAML文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            YAML数据字典
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    
    @staticmethod
    def write_yaml(file_path: str, data: Dict[str, Any]):
        """
        写入YAML文件
        
        Args:
            file_path: 文件路径
            data: 数据字典
        """
        # 确保目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, allow_unicode=True, default_flow_style=False)
    
    @staticmethod
    def read_json(file_path: str) -> Dict[str, Any]:
        """
        读取JSON文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            JSON数据字典
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    @staticmethod
    def write_json(file_path: str, data: Dict[str, Any], indent: int = 2):
        """
        写入JSON文件
        
        Args:
            file_path: 文件路径
            data: 数据字典
            indent: 缩进空格数
        """
        # 确保目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
    
    @staticmethod
    def read_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        读取Excel文件
        
        Args:
            file_path: 文件路径
            sheet_name: 工作表名称，为None时读取第一个工作表
            
        Returns:
            数据列表，每行是一个字典
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("读取Excel需要安装openpyxl: pip install openpyxl")
        workbook = load_workbook(file_path, data_only=True)
        
        if sheet_name is None:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
        
        # 获取表头
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # 读取数据
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
        
        workbook.close()
        return data
    
    @staticmethod
    def write_excel(file_path: str, data: List[Dict[str, Any]],
                    sheet_name: str = 'Sheet1'):
        """
        写入Excel文件

        Args:
            file_path: 文件路径
            data: 数据列表，每行是一个字典
            sheet_name: 工作表名称
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("写入Excel需要安装openpyxl: pip install openpyxl")
        # 确保目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        if not data:
            workbook.save(file_path)
            return
        
        # 写入表头
        headers = list(data[0].keys())
        sheet.append(headers)
        
        # 写入数据
        for row_data in data:
            row = [row_data.get(header, '') for header in headers]
            sheet.append(row)
        
        workbook.save(file_path)
        workbook.close()
    
    @staticmethod
    def read_csv(file_path: str, encoding: str = 'utf-8') -> List[Dict[str, Any]]:
        """
        读取CSV文件
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            数据列表，每行是一个字典
        """
        data = []
        with open(file_path, 'r', encoding=encoding, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(dict(row))
        return data
    
    @staticmethod
    def write_csv(file_path: str, data: List[Dict[str, Any]], 
                  encoding: str = 'utf-8'):
        """
        写入CSV文件
        
        Args:
            file_path: 文件路径
            data: 数据列表，每行是一个字典
            encoding: 文件编码
        """
        # 确保目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        if not data:
            return
        
        headers = list(data[0].keys())
        
        with open(file_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
    
    @staticmethod
    def ensure_dir(dir_path: str):
        """
        确保目录存在
        
        Args:
            dir_path: 目录路径
        """
        os.makedirs(dir_path, exist_ok=True)
    
    @staticmethod
    def get_project_root() -> str:
        """
        获取项目根目录
        
        Returns:
            项目根目录路径
        """
        current_file = Path(__file__).resolve()
        return str(current_file.parent.parent)
    
    @staticmethod
    def get_data_dir() -> str:
        """
        获取数据目录
        
        Returns:
            数据目录路径
        """
        return os.path.join(FileUtils.get_project_root(), 'data')
    
    @staticmethod
    def get_report_dir() -> str:
        """
        获取报告目录
        
        Returns:
            报告目录路径
        """
        return os.path.join(FileUtils.get_project_root(), 'reports')
    
    @staticmethod
    def save_screenshot(page, name: str = None) -> str:
        """
        保存截图
        
        Args:
            page: Playwright页面对象
            name: 截图名称
            
        Returns:
            截图文件路径
        """
        from datetime import datetime
        
        if name is None:
            name = f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        screenshot_dir = os.path.join(FileUtils.get_report_dir(), 'screenshots')
        FileUtils.ensure_dir(screenshot_dir)
        
        screenshot_path = os.path.join(screenshot_dir, f"{name}.png")
        page.screenshot(path=screenshot_path, full_page=True)
        
        return screenshot_path
    
    @staticmethod
    def save_page_source(page, name: str = None) -> str:
        """
        保存页面源码
        
        Args:
            page: Playwright页面对象
            name: 文件名称
            
        Returns:
            文件路径
        """
        from datetime import datetime
        
        if name is None:
            name = f"page_source_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        report_dir = FileUtils.get_report_dir()
        FileUtils.ensure_dir(report_dir)
        
        file_path = os.path.join(report_dir, f"{name}.html")
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(page.content())
        
        return file_path
